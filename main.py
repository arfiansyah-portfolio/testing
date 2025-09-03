import os
import json
import glob
import subprocess
import shutil
import concurrent.futures
from multiprocessing import cpu_count
from tqdm import tqdm
from openpyxl import load_workbook, Workbook

LOCAL_XLSX_DIR = "./datatest/"
TMP_DIR = "./tmp"


def excel_to_json_streaming(xlsx_path, json_path):
    """Convert Excel to JSON with memory optimization"""
    try:
        # Use read_only mode for better memory efficiency
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        data = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            headers = None

            # Stream rows instead of loading all at once
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h) if h else "" for h in row]
                else:
                    if any(cell is not None and str(cell).strip() for cell in row):
                        # Convert cells to strings and handle None values
                        processed_row = [str(cell) if cell is not None else "" for cell in row]
                        rows.append(processed_row)

            data[sheet_name] = {
                "headers": headers or [],
                "rows": rows,
            }

        wb.close()  # Important: free memory immediately

        # Write JSON with minimal memory footprint
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return True

    except Exception as e:
        print(f"Error processing {xlsx_path}: {e}")
        return False


def json_to_excel_optimized(json_path, xlsx_path):
    """Convert JSON back to Excel with memory optimization"""
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        wb = Workbook()

        # Remove default sheet if we have custom sheets
        if len(data) > 0:
            wb.remove(wb.active)

        for sheet_name, content in data.items():
            ws = wb.create_sheet(title=sheet_name)
            headers = content.get("headers", [])
            rows = content.get("rows", [])

            # Write headers
            if headers:
                ws.append(headers)

            # Write data rows in batches to manage memory
            batch_size = 100
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                for row in batch:
                    ws.append(row)

        wb.save(xlsx_path)
        wb.close()
        return True

    except Exception as e:
        print(f"Error converting {json_path} to Excel: {e}")
        return False


def process_file_conversion(file_info):
    """Convert single Excel file (for parallel execution)"""
    file_path, target_dir = file_info
    name = os.path.basename(file_path).replace(".xlsx", ".json")
    json_path = f"{target_dir}/{name}"

    success = excel_to_json_streaming(file_path, json_path)
    return (name, success)


def run_git_pull(branch="automation-epc-ui"):
    """Checkout and pull latest branch with safety checks"""
    try:
        # Check for uncommitted changes
        result = subprocess.run(
            ["git", "status", "--porcelain"],
            capture_output=True, text=True, check=True
        )
        if result.stdout.strip():
            print("❌ Error: You have uncommitted changes. Please commit or stash first.")
            print("Run: git status")
            return False

        # Check current branch
        current_branch = subprocess.run(
            ["git", "branch", "--show-current"],
            capture_output=True, text=True, check=True
        )

        if current_branch.stdout.strip() != branch:
            print(f"🔄 Switching to branch: {branch}")
            subprocess.run(["git", "checkout", branch], check=True)

        print(f"🔄 Pulling latest changes from {branch}...")
        subprocess.run(["git", "pull", "origin", branch], check=True)
        return True

    except subprocess.CalledProcessError as e:
        print(f"❌ Git operation failed: {e}")
        return False


def get_user_choice(prompt, valid_choices):
    """Get validated user input"""
    while True:
        choice = input(f"{prompt} [{'/'.join(valid_choices)}]: ").strip()
        if choice in valid_choices:
            return choice
        print(f"Invalid choice. Please enter one of: {', '.join(valid_choices)}")


def display_conflict_detailed(sheet, row_idx, headers, lrow, rrow):
    """Display conflict with better formatting"""
    print(f"\n{'=' * 60}")
    print(f"CONFLICT in sheet '{sheet}', row {row_idx + 2}")
    print(f"{'=' * 60}")

    # Show side-by-side comparison
    max_header_len = max(len(h) for h in headers) if headers else 10

    if lrow:
        print("LOCAL VALUES:")
        for i, (header, value) in enumerate(zip(headers, lrow)):
            if i < len(headers):
                print(f"  {header:<{max_header_len}}: {value}")

    if rrow:
        print("\nREMOTE VALUES:")
        for i, (header, value) in enumerate(zip(headers, rrow)):
            if i < len(headers):
                print(f"  {header:<{max_header_len}}: {value}")


def merge_rows_enhanced(headers, lrow, rrow):
    """Enhanced row merging with better conflict resolution"""
    if not lrow and not rrow:
        return None
    if not lrow:
        return list(rrow)
    if not rrow:
        return list(lrow)

    merged = []
    conflicts_found = False

    max_len = max(len(lrow), len(rrow))

    for i in range(max_len):
        lv = lrow[i] if i < len(lrow) else ""
        rv = rrow[i] if i < len(rrow) else ""
        header = headers[i] if i < len(headers) else f"Column_{i + 1}"

        # Clean empty values
        lv_clean = str(lv).strip() if lv is not None else ""
        rv_clean = str(rv).strip() if rv is not None else ""

        if lv_clean == rv_clean:
            merged.append(lv_clean)
        elif lv_clean and not rv_clean:
            merged.append(lv_clean)
        elif rv_clean and not lv_clean:
            merged.append(rv_clean)
        else:
            # Real conflict - different non-empty values
            conflicts_found = True
            print(f"\n🔥 Field conflict in '{header}':")
            print(f"   Local : '{lv_clean}'")
            print(f"   Remote: '{rv_clean}'")
            choice = get_user_choice("Choose", ["1", "2"])
            merged.append(lv_clean if choice == "1" else rv_clean)

    return merged


def compare_json_enhanced(local_file, remote_file, resolved_file):
    """Enhanced JSON comparison with better conflict handling"""
    print(f"🔍 Comparing {os.path.basename(local_file)}...")

    try:
        with open(local_file, "r", encoding="utf-8") as f:
            local_data = json.load(f)
        with open(remote_file, "r", encoding="utf-8") as f:
            remote_data = json.load(f)
    except Exception as e:
        print(f"❌ Error reading JSON files: {e}")
        return False

    resolved = {}
    total_conflicts = 0

    all_sheets = set(local_data.keys()).union(remote_data.keys())

    for sheet in all_sheets:
        print(f"  📋 Processing sheet: {sheet}")

        local_sheet = local_data.get(sheet, {"headers": [], "rows": []})
        remote_sheet = remote_data.get(sheet, {"headers": [], "rows": []})

        headers = local_sheet.get("headers", []) or remote_sheet.get("headers", [])
        local_rows = local_sheet.get("rows", [])
        remote_rows = remote_sheet.get("rows", [])

        max_len = max(len(local_rows), len(remote_rows))
        resolved_rows = []
        sheet_conflicts = 0

        for i in range(max_len):
            lrow = local_rows[i] if i < len(local_rows) else None
            rrow = remote_rows[i] if i < len(remote_rows) else None

            if lrow == rrow:
                if lrow and any(str(cell).strip() for cell in lrow):
                    resolved_rows.append(lrow)
            else:
                sheet_conflicts += 1
                display_conflict_detailed(sheet, i, headers, lrow, rrow)

                choice = get_user_choice(
                    "\nResolve conflict",
                    ["1", "2", "3", "4", "s"]  # s for skip
                )

                if choice == "1" and lrow:
                    resolved_rows.append(lrow)
                elif choice == "2" and rrow:
                    resolved_rows.append(rrow)
                elif choice == "3":
                    merged = merge_rows_enhanced(headers, lrow, rrow)
                    if merged:
                        resolved_rows.append(merged)
                elif choice == "4":
                    merged = merge_rows_enhanced(headers, rrow, lrow)
                    if merged:
                        resolved_rows.append(merged)
                # choice == "s" means skip row

        resolved[sheet] = {
            "headers": headers,
            "rows": [r for r in resolved_rows if r and any(str(cell).strip() for cell in r)],
        }

        if sheet_conflicts > 0:
            print(f"  ✅ Resolved {sheet_conflicts} conflicts in sheet '{sheet}'")

        total_conflicts += sheet_conflicts

    # Save resolved JSON
    with open(resolved_file, "w", encoding="utf-8") as f:
        json.dump(resolved, f, indent=2, ensure_ascii=False)

    print(f"✅ Total conflicts resolved: {total_conflicts}")
    return True


def main():
    """Main function with enhanced error handling and progress tracking"""
    print("🚀 Starting Enhanced Excel Merge Process...")
    print("=" * 50)

    # Create directories
    dirs_to_create = [
        f"{TMP_DIR}/local_json",
        f"{TMP_DIR}/remote_json",
        f"{TMP_DIR}/resolved_json"
    ]

    for dir_path in dirs_to_create:
        os.makedirs(dir_path, exist_ok=True)

    # Get all Excel files
    excel_files = glob.glob(f"{LOCAL_XLSX_DIR}/*.xlsx")
    if not excel_files:
        print("❌ No Excel files found in", LOCAL_XLSX_DIR)
        return

    print(f"📁 Found {len(excel_files)} Excel files")

    # Step 1: Convert local Excel files to JSON (parallel)
    print("\n📊 Step 1/5: Converting local Excel files to JSON...")
    file_infos = [(f, f"{TMP_DIR}/local_json") for f in excel_files]

    max_workers = min(4, cpu_count(), len(excel_files))

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_file_conversion, info) for info in file_infos]

        successful_local = []
        with tqdm(total=len(futures), desc="Converting local files") as pbar:
            for future in concurrent.futures.as_completed(futures):
                filename, success = future.result()
                if success:
                    successful_local.append(filename)
                pbar.update(1)

    print(f"✅ Successfully converted {len(successful_local)}/{len(excel_files)} local files")

    # Step 2: Git pull
    print("\n🌐 Step 2/5: Pulling latest changes from Git...")
    if not run_git_pull():
        print("❌ Git pull failed. Aborting merge process.")
        return

    # Step 3: Convert remote Excel files to JSON (parallel)
    print("\n📊 Step 3/5: Converting remote Excel files to JSON...")
    remote_excel_files = glob.glob(f"{LOCAL_XLSX_DIR}/*.xlsx")
    file_infos = [(f, f"{TMP_DIR}/remote_json") for f in remote_excel_files]

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_file_conversion, info) for info in file_infos]

        successful_remote = []
        with tqdm(total=len(futures), desc="Converting remote files") as pbar:
            for future in concurrent.futures.as_completed(futures):
                filename, success = future.result()
                if success:
                    successful_remote.append(filename)
                pbar.update(1)

    print(f"✅ Successfully converted {len(successful_remote)}/{len(remote_excel_files)} remote files")

    # Step 4: Compare and resolve conflicts
    print("\n🔍 Step 4/5: Comparing files and resolving conflicts...")

    files_to_compare = set(successful_local).intersection(set(successful_remote))
    print(f"📋 Comparing {len(files_to_compare)} files...")

    resolved_files = []
    for filename in tqdm(files_to_compare, desc="Resolving conflicts"):
        local_file = f"{TMP_DIR}/local_json/{filename}"
        remote_file = f"{TMP_DIR}/remote_json/{filename}"
        resolved_file = f"{TMP_DIR}/resolved_json/{filename}"

        if compare_json_enhanced(local_file, remote_file, resolved_file):
            resolved_files.append(filename)

    # Handle files that exist only locally or remotely
    only_local = set(successful_local) - set(successful_remote)
    only_remote = set(successful_remote) - set(successful_local)

    for filename in only_local:
        print(f"📝 File {filename} only exists locally - keeping local version")
        shutil.copy(f"{TMP_DIR}/local_json/{filename}", f"{TMP_DIR}/resolved_json/{filename}")
        resolved_files.append(filename)

    for filename in only_remote:
        print(f"📝 File {filename} only exists remotely - using remote version")
        shutil.copy(f"{TMP_DIR}/remote_json/{filename}", f"{TMP_DIR}/resolved_json/{filename}")
        resolved_files.append(filename)

    # Step 5: Convert resolved JSON back to Excel
    print(f"\n💾 Step 5/5: Converting {len(resolved_files)} resolved files back to Excel...")

    conversion_success = 0
    for filename in tqdm(resolved_files, desc="Converting to Excel"):
        json_file = f"{TMP_DIR}/resolved_json/{filename}"
        xlsx_file = f"{LOCAL_XLSX_DIR}/{filename.replace('.json', '.xlsx')}"

        if json_to_excel_optimized(json_file, xlsx_file):
            conversion_success += 1

    print("\n" + "=" * 50)
    print("🎉 MERGE PROCESS COMPLETED!")
    print(f"✅ Successfully processed: {conversion_success}/{len(resolved_files)} files")

    # Cleanup option
    cleanup = get_user_choice("\n🧹 Clean up temporary files?", ["y", "n"])
    if cleanup == "y":
        shutil.rmtree(TMP_DIR)
        print("🗑️  Temporary files cleaned up")

    print("🏁 All done!")


if __name__ == "__main__":
    main()